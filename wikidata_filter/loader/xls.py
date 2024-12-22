from datetime import datetime
from wikidata_filter.loader.file import File

DATETIME_FORMAT = '%Y-%m-%d %H:%M'

try:
    import openpyxl
except ImportError:
    raise ImportError("failed to import openpyxl  needed by Excel parsing")


class ExcelStream(File):
    """Excel流式读取，适合大文件"""
    def __init__(self, input_file: str, sheets: list = None, with_header: bool = True):
        self.input_file = input_file
        self.sheets = sheets
        self.with_header = with_header

        self.instream = openpyxl.load_workbook(self.input_file, read_only=True)

        sheet_names = self.instream.sheetnames
        if self.sheets:
            sheet_names = [i if isinstance(i, str) else sheet_names[i] for i in self.sheets]
        # print("sheet_names", sheet_names)
        self.sheet_names = sheet_names

    def parse_row(self, row: tuple):
        vals = []
        for v in row:
            if isinstance(v, datetime):
                vals.append(v.strftime(DATETIME_FORMAT))
            else:
                vals.append(v)

        return vals

    def iter(self):
        for sheet_name in self.sheet_names:
            sheet = self.instream.get_sheet_by_name(sheet_name)

            header = None
            for row in sheet.iter_rows(values_only=True):
                row_values = self.parse_row(row)
                if self.with_header:
                    if header is None:
                        header = row_values
                    else:
                        output = dict(zip(header, row_values))
                        output['_sheet'] = sheet_name
                        yield output
                else:
                    yield sheet_name, row_values
